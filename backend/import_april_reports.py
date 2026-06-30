"""
Script Import Data Laporan Harian Unspec Bulan April 2026
=========================================================
Membaca file Excel harian dari folder 'data' di root project,
mengekstrak total saldo, close, dan saldo lama dari sheet "data unspec HI",
lalu insert/update ke tabel daily_reports (PostgreSQL/SQLite).

Cara jalankan:
  python backend/import_april_reports.py
"""

import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl belum terinstall. Jalankan: pip install openpyxl")
    sys.exit(1)

from sqlalchemy import create_engine, Column, Integer, Date, TIMESTAMP, func, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

# === KONFIGURASI DATABASE ===
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://gpon_user:gpon_secure_pass_2026@localhost:5432/gpon_network"
)

# === SETUP PATH ===
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "data"

# Target default saldo unspec
TARGET_DEFAULT = 127

# === SETUP DATABASE ===
Base = declarative_base()

class DailyReportDB(Base):
    """Model tabel daily_reports"""
    __tablename__ = "daily_reports"
    date = Column(Date, primary_key=True, index=True)
    total_saldo = Column(Integer, default=0)
    close = Column(Integer, default=0)
    saldo_lama = Column(Integer, default=0)
    target = Column(Integer, default=0)
    created_at = Column(TIMESTAMP(timezone=True), server_default=func.now())
    updated_at = Column(TIMESTAMP(timezone=True), server_default=func.now(), onupdate=func.now())


# === PETA BULAN INDONESIA ===
BULAN_MAP = {
    "JANUARI": 1,   "FEBRUARI": 2,  "MARET": 3,
    "APRIL": 4,     "MEI": 5,       "JUNI": 6,
    "JULI": 7,      "AGUSTUS": 8,   "SEPTEMBER": 9,
    "OKTOBER": 10,  "NOVEMBER": 11, "DESEMBER": 12,
}


def ekstrak_tanggal_dari_nama_file(nama_file: str):
    """
    Ekstrak tanggal dari nama file Excel.
    Contoh: 'LAPORAN UNSPEC HD 14 APRIL 2026 (2).xlsx' -> date(2026, 4, 14)
    """
    try:
        pola = r"(?:HD|TEKNISI).*?(\d{1,2})\s+(\w+)\s+(\d{4})"
        cocok = re.search(pola, nama_file, re.IGNORECASE)
        if not cocok:
            return None

        tanggal = int(cocok.group(1))
        bulan_str = cocok.group(2).upper()
        tahun = int(cocok.group(3))

        bulan = BULAN_MAP.get(bulan_str)
        if not bulan:
            print(f"  [!] Bulan '{bulan_str}' tidak dikenali")
            return None

        return date(tahun, bulan, tanggal)
    except Exception as e:
        print(f"  [!] Gagal parse tanggal dari nama file: {e}")
        return None



def cari_sheet_hi(workbook):
    """Mencari sheet 'data unspec HI' atau yang serupa"""
    for nama_sheet in workbook.sheetnames:
        nama_upper = nama_sheet.upper()
        if "HI" in nama_upper or "UNSPEC" in nama_upper:
            return nama_sheet

    # Fallback: ambil sheet pertama yang bukan sheet bawaan
    for nama_sheet in workbook.sheetnames:
        if nama_sheet.upper() not in ["KURMA", "SHEET1"]:
            return nama_sheet
    return None


def hitung_saldo_dari_sheet(file_path: str):
    """
    Membaca sheet data unspec, menghitung total saldo, close, dan saldo lama.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        nama_sheet = cari_sheet_hi(wb)
        if not nama_sheet:
            print(f"  [!] Sheet data unspec tidak ditemukan")
            wb.close()
            return None

        ws = wb[nama_sheet]

        # Cari baris header
        baris_header = None
        header_cols = {}

        for baris_idx, baris in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
            for sel in baris:
                if sel.value and str(sel.value).strip().upper() == "NO":
                    baris_header = baris_idx
                    for col_sel in baris:
                        if col_sel.value:
                            header_cols[str(col_sel.value).strip().upper()] = col_sel.column - 1
                    break
            if baris_header:
                break

        if not baris_header:
            print(f"  [!] Header row tidak ditemukan")
            wb.close()
            return None

        # Cari kolom KET (CLOSED/SISA TIKET)
        idx_ket = header_cols.get("KET")

        if idx_ket is None:
            # Fallback: scan baris pertama data untuk mencari kolom KET
            for baris in ws.iter_rows(min_row=baris_header + 1, max_row=baris_header + 1, values_only=False):
                for sel in baris:
                    val = str(sel.value).upper() if sel.value else ""
                    if val in ["CLOSED", "SISA TIKET"]:
                        idx_ket = sel.column - 1
                        break
                break

        if idx_ket is None:
            print(f"  [!] Kolom KET (CLOSED/SISA TIKET) tidak ditemukan")
            wb.close()
            return None

        # Hitung data baris per baris
        total_saldo = 0
        total_close = 0

        for baris in ws.iter_rows(min_row=baris_header + 1, values_only=True):
            # Lewati baris kosong
            if not baris[0] or str(baris[0]).strip() == "":
                continue

            # Validasi kolom No harus berupa angka integer
            try:
                int(str(baris[0]).strip())
            except ValueError:
                continue

            total_saldo += 1

            # Cek status closed di kolom KET
            ket_val = str(baris[idx_ket]).strip().upper() if len(baris) > idx_ket and baris[idx_ket] else ""
            if ket_val == "CLOSED":
                total_close += 1

        wb.close()

        return {
            "total_saldo": total_saldo,
            "close": total_close,
            "saldo_lama": total_saldo - total_close,
        }

    except Exception as e:
        print(f"  [X] Error saat memproses file: {e}")
        return None


def sort_key_file(file_path):
    """
    Mengurutkan berdasarkan tanggal, lalu file revisi (2) ditaruh terakhir 
    agar menimpa file tanpa suffix jika tanggalnya sama.
    """
    nama_file = file_path.name
    tanggal = ekstrak_tanggal_dari_nama_file(nama_file)
    if not tanggal:
        return (date(1970, 1, 1), 0, nama_file)
    
    # Deteksi suffix revisi (2)
    has_suffix = 1 if "(2)" in nama_file else 0
    return (tanggal, has_suffix, nama_file)


def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print("  IMPORT DATA LAPORAN BULAN APRIL 2026")
    if dry_run:
        print("  MODE: DRY-RUN (Hanya simulasi parsing)")
    print("=" * 60)

    session = None

    if not dry_run:
        # Koneksi database dengan fallback cerdas
        try:
            print(f"Mencoba koneksi ke PostgreSQL: {DATABASE_URL}")
            engine = create_engine(DATABASE_URL)
            Session = sessionmaker(bind=engine)
            session = Session()
            session.execute(text("SELECT 1"))
            print("[OK] Koneksi PostgreSQL berhasil\n")
        except Exception as pg_err:
            print(f"[!] Gagal konek PostgreSQL: {pg_err}")
            
            # Fallback ke SQLite lokal
            sqlite_path = SCRIPT_DIR / "gpon_network.db"
            if not sqlite_path.exists():
                sqlite_path = SCRIPT_DIR.parent / "backend" / "gpon_network.db"
            
            sqlite_url = f"sqlite:///{sqlite_path}"
            print(f"Mencoba fallback ke SQLite lokal: {sqlite_url}")
            try:
                engine = create_engine(sqlite_url)
                Session = sessionmaker(bind=engine)
                session = Session()
                session.execute(text("SELECT 1"))
                print("[OK] Koneksi SQLite lokal berhasil\n")
            except Exception as sqlite_err:
                print(f"[GAGAL] Koneksi SQLite: {sqlite_err}")
                sys.exit(1)
    else:
        print("[INFO] Mode dry-run: skip koneksi database\n")

    # Kumpulkan semua file Excel di folder data
    if not DATA_DIR.exists():
        print(f"[GAGAL] Folder data tidak ditemukan di: {DATA_DIR}")
        if session:
            session.close()
        sys.exit(1)

    file_xlsx = [f for f in DATA_DIR.glob("*.xlsx") if not f.name.startswith("~$")]
    
    # Filter khusus bulan April
    file_april = []
    for f in file_xlsx:
        tgl = ekstrak_tanggal_dari_nama_file(f.name)
        if tgl and tgl.month == 4 and tgl.year == 2026:
            file_april.append(f)

    # Urutkan file dengan sort_key_file
    file_april.sort(key=sort_key_file)

    print(f"[INFO] Menemukan {len(file_april)} file laporan bulan April 2026 di folder data")
    
    if not file_april:
        print("[!] Tidak ada file bulan April 2026 yang terdeteksi.")
        if session:
            session.close()
        sys.exit(0)

    print("-" * 60)

    berhasil = 0
    gagal = 0
    hasil_semua = []

    for file_path in file_april:
        nama_file = file_path.name
        print(f"\n[FILE] {nama_file}")

        tanggal = ekstrak_tanggal_dari_nama_file(nama_file)
        if not tanggal:
            print(f"  [!] Skip - Tanggal tidak valid")
            gagal += 1
            continue

        print(f"  Tanggal Terdeteksi: {tanggal.isoformat()}")

        # Ekstrak data
        hasil = hitung_saldo_dari_sheet(str(file_path))
        if not hasil:
            gagal += 1
            continue

        print(f"  Total Saldo       : {hasil['total_saldo']}")
        print(f"  Total Close       : {hasil['close']}")
        # Saldo lama = total_saldo - close
        print(f"  Saldo Lama        : {hasil['saldo_lama']}")

        hasil_semua.append({
            "tanggal": tanggal,
            **hasil
        })

        if dry_run:
            berhasil += 1
            continue

        # Simpan atau update ke database
        try:
            report = session.query(DailyReportDB).filter(
                DailyReportDB.date == tanggal
            ).first()

            if report:
                report.total_saldo = hasil["total_saldo"]
                report.close = hasil["close"]
                report.saldo_lama = hasil["saldo_lama"]
                report.target = TARGET_DEFAULT
                print(f"  [UPDATE] Data tanggal {tanggal} diperbarui")
            else:
                report = DailyReportDB(
                    date=tanggal,
                    total_saldo=hasil["total_saldo"],
                    close=hasil["close"],
                    saldo_lama=hasil["saldo_lama"],
                    target=TARGET_DEFAULT,
                )
                session.add(report)
                print(f"  [INSERT] Data tanggal {tanggal} dimasukkan baru")

            session.commit()
            berhasil += 1
        except Exception as db_err:
            session.rollback()
            print(f"  [X] Gagal menyimpan ke database: {db_err}")
            gagal += 1

    # Print Ringkasan
    print("\n" + "=" * 60)
    print("  RINGKASAN IMPORT BULAN APRIL")
    print("=" * 60)
    print(f"  Berhasil diproses : {berhasil}")
    print(f"  Gagal diproses    : {gagal}")
    print(f"  Total File April  : {len(file_april)}")

    if hasil_semua:
        print("\n" + "-" * 60)
        print(f"  {'Tanggal':<14} {'Total':>6} {'Close':>6} {'Saldo':>6}")
        print(f"  {'-'*14} {'-'*6} {'-'*6} {'-'*6}")
        for h in sorted(hasil_semua, key=lambda x: x["tanggal"]):
            print(f"  {h['tanggal'].isoformat():<14} {h['total_saldo']:>6} {h['close']:>6} {h['saldo_lama']:>6}")

    if session:
        session.close()

    print("\n[SELESAI]")


if __name__ == "__main__":
    main()
