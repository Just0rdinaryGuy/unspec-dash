import openpyxl
import os

# Cek sheet "data unspec HI" di file FEB dan MAR
files = [
    r"d:\code\unspec-dash-main\FEB 2026\LAPORAN UNSPEC HD 01 FEBRUARI 2026.xlsx",
    r"d:\code\unspec-dash-main\MAR 2026\LAPORAN UNSPEC HD 03 MARET 2026.xlsx",
]

for f in files:
    print(f"\n{'='*80}")
    print(f"FILE: {os.path.basename(f)}")
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"ALL SHEETS: {wb.sheetnames}")
    
    # Cari sheet yang mengandung 'HI' atau 'unspec'
    for sn in wb.sheetnames:
        sn_upper = sn.upper()
        if 'HI' in sn_upper or 'UNSPEC' in sn_upper or 'DATA' in sn_upper:
            ws = wb[sn]
            print(f"\n--- Sheet: '{sn}' (rows={ws.max_row}, cols={ws.max_column}) ---")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False), 1):
                vals = [str(c.value)[:50] if c.value else "" for c in row[:36]]
                print(f"  Row {row_idx}: {vals}")
            # Juga cek baris terakhir (total?)
            print(f"\n  --- Last 5 rows ---")
            for row_idx, row in enumerate(ws.iter_rows(min_row=max(1, ws.max_row-4), max_row=ws.max_row, values_only=False), max(1, ws.max_row-4)):
                vals = [str(c.value)[:50] if c.value else "" for c in row[:36]]
                print(f"  Row {row_idx}: {vals}")
    
    # Juga cek sheet KURMA summary
    for sn in wb.sheetnames:
        if 'KURMA' in sn.upper():
            ws = wb[sn]
            print(f"\n--- Sheet: '{sn}' (rows={ws.max_row}, cols={ws.max_column}) ---")
            # Last rows
            print(f"  --- Last 3 rows ---")
            for row_idx, row in enumerate(ws.iter_rows(min_row=max(1, ws.max_row-2), max_row=ws.max_row, values_only=False), max(1, ws.max_row-2)):
                vals = [str(c.value)[:50] if c.value else "" for c in row[:10]]
                print(f"  Row {row_idx}: {vals}")
    
    wb.close()
