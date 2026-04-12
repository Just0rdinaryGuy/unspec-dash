import openpyxl

# Debug: cek kolom STATUS KURMA di file FEB dan MAR
files = [
    r"d:\code\unspec-dash-main\FEB 2026\LAPORAN UNSPEC HD 01 FEBRUARI 2026.xlsx",
    r"d:\code\unspec-dash-main\MAR 2026\LAPORAN UNSPEC HD 03 MARET 2026.xlsx",
]

for f in files:
    print(f"\n{'='*60}")
    print(f"FILE: {f.split('\\')[-1]}")
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    
    # Cari sheet HI
    target_sheet = None
    for sn in wb.sheetnames:
        sn_upper = sn.upper()
        if "HI" in sn_upper or "UNSPEC" in sn_upper:
            target_sheet = sn
            break
    if not target_sheet:
        for sn in wb.sheetnames:
            if sn.upper() not in ["KURMA", "SHEET1"]:
                target_sheet = sn
                break
    
    print(f"Sheet: '{target_sheet}'")
    ws = wb[target_sheet]
    
    # Cari header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "NO":
                print(f"Header row: {row_idx}")
                # Print all header values with index
                for c in row:
                    if c.value:
                        print(f"  Col {c.column-1} (col_letter={c.column}): '{c.value}'")
                
                # Now print first 3 data rows with STATUS KURMA column
                print(f"\nData rows (col 23=STATUS KURMA, col 24=KET):")
                for data_row in ws.iter_rows(min_row=row_idx+1, max_row=row_idx+5, values_only=True):
                    if data_row[0]:
                        # Print No, ND (col 10), STATUS KURMA (col 23), KET (col 24)
                        no = data_row[0]
                        nd = data_row[10] if len(data_row) > 10 else "?"
                        sk = data_row[23] if len(data_row) > 23 else "?"
                        ket = data_row[24] if len(data_row) > 24 else "?"
                        print(f"  No={no}, ND={nd}, StatusKurma='{sk}', Ket='{ket}'")
                break
        else:
            continue
        break
    
    wb.close()
