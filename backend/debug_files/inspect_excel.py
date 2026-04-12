import openpyxl
import os

# Cek satu file FEB dan satu file MAR
files = [
    r"d:\code\unspec-dash-main\FEB 2026\LAPORAN UNSPEC HD 01 FEBRUARI 2026.xlsx",
    r"d:\code\unspec-dash-main\MAR 2026\LAPORAN UNSPEC HD 03 MARET 2026.xlsx",
]

for f in files:
    print(f"\n{'='*60}")
    print(f"FILE: {os.path.basename(f)}")
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"SHEETS: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' (rows={ws.max_row}, cols={ws.max_column}) ---")
        # Print header + 5 baris pertama
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=False), 1):
            vals = [str(c.value)[:40] if c.value else "" for c in row]
            print(f"  Row {row_idx}: {vals}")
    wb.close()
