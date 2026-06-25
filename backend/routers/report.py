from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import date, datetime as dt, timedelta, time
import re
import openpyxl
import tempfile
import os
from database import get_db, NetworkNodeDB
from models.report import DailyReportDB, DailyReportResponse, DailyReportUpdate

BULAN_MAP = {
    "JANUARI": 1,   "FEBRUARI": 2,  "MARET": 3,
    "APRIL": 4,     "MEI": 5,       "JUNI": 6,
    "JULI": 7,      "AGUSTUS": 8,   "SEPTEMBER": 9,
    "OKTOBER": 10,  "NOVEMBER": 11, "DESEMBER": 12,
}

def ekstrak_tanggal_dari_nama_file(nama_file: str):
    try:
        pola = r"HD\s+(\d{1,2})\s+(\w+)\s+(\d{4})"
        cocok = re.search(pola, nama_file, re.IGNORECASE)
        if not cocok:
            return None

        tanggal = int(cocok.group(1))
        bulan_str = cocok.group(2).upper()
        tahun = int(cocok.group(3))

        bulan = BULAN_MAP.get(bulan_str)
        if not bulan:
            return None

        return date(tahun, bulan, tanggal)
    except:
        return None

def cari_sheet_hi(workbook):
    for nama_sheet in workbook.sheetnames:
        nama_upper = nama_sheet.upper()
        if "HI" in nama_upper or "UNSPEC" in nama_upper:
            return nama_sheet
    for nama_sheet in workbook.sheetnames:
        if nama_sheet.upper() not in ["KURMA", "SHEET1"]:
            return nama_sheet
    return None

def hitung_saldo_dari_sheet(file_path: str):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        nama_sheet = cari_sheet_hi(wb)
        if not nama_sheet:
            wb.close()
            return None

        ws = wb[nama_sheet]
        baris_header = None
        header_cols = {}

        for baris_idx, baris in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
            for sel in baris:
                if sel.value and str(sel.value).strip().upper() == "NO":
                    baris_header = baris_idx
                    for col_sel in baris:
                        if col_sel.value:
                            header_cols[str(col_sel.value).strip().upper()] = col_sel.column - 1
                    break
            if baris_header:
                break

        if not baris_header:
            wb.close()
            return None

        idx_ket = header_cols.get("KET")
        if idx_ket is None:
            for baris in ws.iter_rows(min_row=baris_header + 1, max_row=baris_header + 1, values_only=False):
                for sel in baris:
                    val = str(sel.value).upper() if sel.value else ""
                    if val in ["CLOSED", "SISA TIKET"]:
                        idx_ket = sel.column - 1
                        break
                break

        if idx_ket is None:
            wb.close()
            return None

        total_saldo = 0
        total_close = 0

        for baris in ws.iter_rows(min_row=baris_header + 1, values_only=True):
            if not baris[0] or str(baris[0]).strip() == "":
                continue
            try:
                int(str(baris[0]).strip())
            except ValueError:
                continue

            total_saldo += 1
            ket_val = str(baris[idx_ket]).strip().upper() if len(baris) > idx_ket and baris[idx_ket] else ""
            if ket_val == "CLOSED":
                total_close += 1

        wb.close()
        return {
            "total_saldo": total_saldo,
            "close": total_close,
            "saldo_lama": total_saldo - total_close,
        }
    except:
        return None

from middleware.auth_middleware import get_current_active_user

router = APIRouter(
    prefix="/api/reports",
    tags=["Reports"],
    dependencies=[Depends(get_current_active_user)]
)

@router.get("/", response_model=List[DailyReportResponse])
async def get_reports(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db)
):
    """
    Ambil laporan harian buat bulan dan tahun tertentu
    """
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
        
    reports = db.query(DailyReportDB).filter(
        DailyReportDB.date >= start_date,
        DailyReportDB.date < end_date
    ).order_by(DailyReportDB.date).all()
    
    return reports

@router.post("/generate", response_model=DailyReportResponse)
async def generate_today_report(
    target: int = Query(127, description="Target daily unspec"),
    report_date: Optional[date] = Query(None, description="Date to generate report for (default: Today)"),
    db: Session = Depends(get_db)
):
    """
    Generate atau update report buat HARI INI (atau tanggal tertentu).
    - Total Saldo: Total semua node saat ini
    - Close: Jumlah node CLOSED
    - Saldo Lama: Total - Close (sisa yang belum closed)
    """
    today = report_date if report_date else date.today()
    
    from services.real_data_service import RealDataService
    service = RealDataService(db)
    
    try:
        return service.generate_daily_report(today, target)
    except Exception as e:
        print(f"ERROR GENERATING REPORT: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")

@router.post("/regenerate-all")
async def regenerate_all_reports(
    target: int = Query(127, description="Target daily unspec"),
    db: Session = Depends(get_db)
):
    """
    Regenerate laporan harian untuk SEMUA tanggal yang ada di database.
    Ini berguna kalau ada data yang diimport tapi reportnya belum dibuat.
    """
    from services.real_data_service import RealDataService
    service = RealDataService(db)
    
    try:
        # Get all unique dates from network_nodes table
        distinct_dates = db.query(
            func.date(NetworkNodeDB.import_date).label('date')
        ).distinct().all()
        
        generated_count = 0
        reports = []
        
        for (date_obj,) in distinct_dates:
            if date_obj:
                report = service.generate_daily_report(date_obj, target)
                reports.append({
                    "date": date_obj.isoformat(),
                    "total_saldo": report.total_saldo,
                    "close": report.close,
                    "saldo_lama": report.saldo_lama
                })
                generated_count += 1
        
        return {
            "status": "success",
            "message": f"Regenerated {generated_count} daily reports",
            "reports": reports
        }
        
    except Exception as e:
        print(f"ERROR REGENERATING REPORTS: {e}")
        raise HTTPException(status_code=500, detail=f"Error regenerating reports: {str(e)}")

@router.put("/{report_date}", response_model=DailyReportResponse)
async def update_report(
    report_date: date,
    update_data: DailyReportUpdate,
    db: Session = Depends(get_db)
):
    """
    Manually update a report entry (e.g. to fix Target or adjust numbers)
    """
    report = db.query(DailyReportDB).filter(DailyReportDB.date == report_date).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
        
    for key, value in update_data.dict(exclude_unset=True).items():
        setattr(report, key, value)
        
    db.commit()
    db.refresh(report)
    return report

from fastapi.responses import StreamingResponse

@router.get("/export")
async def export_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000, le=2100),
    db: Session = Depends(get_db)
):
    """
    Export detailed Excel report with Chart and Raw Data Sheets
    """
    from services.real_data_service import RealDataService
    service = RealDataService(db)
    
    try:
        buffer = service.export_monthly_report(month, year)
        
        filename = f"Laporan_Unspec_{year}-{month:02d}.xlsx"
        
        return StreamingResponse(
            buffer, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
            
    except Exception as e:
        print(f"EXPORT ERROR: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.post("/upload-bulk")
async def upload_bulk_reports(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload bulk Excel files laporan harian (HD) dan simpan ke database.
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diunggah")

    TARGET_DEFAULT = 127
    imported_list = []
    errors_list = []

    # Sort files to process revisions (2) last
    def get_sort_key(f):
        tgl = ekstrak_tanggal_dari_nama_file(f.filename)
        has_suffix = 1 if "(2)" in f.filename else 0
        return (tgl or date(1970, 1, 1), has_suffix, f.filename)
        
    sorted_files = sorted(files, key=get_sort_key)

    for file in sorted_files:
        if not file.filename.endswith(('.xlsx', '.xls')):
            errors_list.append(f"{file.filename} bukan file Excel")
            continue

        # Simpan file sementara
        suffix = os.path.splitext(file.filename)[1]
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        try:
            content = await file.read()
            temp_file.write(content)
            temp_file.close()

            # Parse tanggal dari nama file
            tanggal = ekstrak_tanggal_dari_nama_file(file.filename)
            if not tanggal:
                errors_list.append(f"{file.filename}: Nama file tidak sesuai format (contoh: LAPORAN UNSPEC HD 1 APRIL 2026.xlsx)")
                continue

            # Hitung data dari sheet
            data_saldo = hitung_saldo_dari_sheet(temp_file.name)
            if not data_saldo:
                errors_list.append(f"{file.filename}: Gagal membaca data sheet unspec")
                continue

            # Update atau Insert ke Database
            report = db.query(DailyReportDB).filter(DailyReportDB.date == tanggal).first()
            if report:
                report.total_saldo = data_saldo["total_saldo"]
                report.close = data_saldo["close"]
                report.saldo_lama = data_saldo["saldo_lama"]
                report.target = TARGET_DEFAULT
                action = "updated"
            else:
                report = DailyReportDB(
                    date=tanggal,
                    total_saldo=data_saldo["total_saldo"],
                    close=data_saldo["close"],
                    saldo_lama=data_saldo["saldo_lama"],
                    target=TARGET_DEFAULT,
                )
                db.add(report)
                action = "inserted"

            db.commit()
            imported_list.append({
                "filename": file.filename,
                "date": tanggal.isoformat(),
                "total_saldo": data_saldo["total_saldo"],
                "close": data_saldo["close"],
                "saldo_lama": data_saldo["saldo_lama"],
                "action": action
            })

        except Exception as e:
            db.rollback()
            errors_list.append(f"{file.filename}: Error {str(e)}")
        finally:
            try:
                os.unlink(temp_file.name)
            except:
                pass

    return {
        "status": "success" if len(imported_list) > 0 else "failed",
        "imported_count": len(imported_list),
        "failed_count": len(errors_list),
        "imported": imported_list,
        "errors": errors_list
    }
