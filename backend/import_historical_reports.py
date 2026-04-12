"""
Script Import Data Historis Report Saldo Unspec
================================================
Membaca file Excel harian dari folder FEB 2026 & MAR 2026,
extract total saldo, close, dan saldo lama dari sheet "data unspec HI",
lalu insert/update ke tabel daily_reports di PostgreSQL.

Cara pakai (di server):
  # Copy folder FEB 2026 dan MAR 2026 ke server dulu
  # lalu jalankan:
  python import_historical_reports.py

  # Dry-run (validasi parsing tanpa database):
  python import_historical_reports.py --dry-run

Prasyarat:
  pip install openpyxl sqlalchemy psycopg2-binary
"""

import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl belum terinstall. Jalankan: pip install openpyxl")
    sys.exit(1)

from sqlalchemy import create_engine, Column, Integer, Date, TIMESTAMP, func, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

# === KONFIGURASI DATABASE ===
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://gpon_user:gpon_secure_pass_2026@localhost:5432/gpon_network"
)

# === KONFIGURASI FOLDER ===
# Coba cari di dalam folder yang sama dulu (kalo di-run di dalam container)
# lalu fallback ke parent folder (kalo run lokal di luar container)
SCRIPT_DIR = Path(__file__).parent

def get_folder_path(folder_name: str) -> Path:
    if (SCRIPT_DIR / folder_name).exists():
        return SCRIPT_DIR / folder_name
    return SCRIPT_DIR.parent / folder_name

DATA_FOLDERS = {
    "FEB 2026": get_folder_path("FEB 2026"),
    "MAR 2026": get_folder_path("MAR 2026"),
}

# Target default saldo unspec
TARGET_DEFAULT = 127

# === SETUP DATABASE ===
Base = declarative_base()

class DailyReportDB(Base):
    """Model tabel daily_reports"""
    __tablename__ = "daily_reports"
    date = Column(Date, primary_key=True, index=True)
    total_saldo = Column(Integer, default=0)
    close = Column(Integer, default=0)
    saldo_lama = Column(Integer, default=0)
    target = Column(Integer, default=0)
    created_at = Column(TIMESTAMP(timezone=True), server_default=func.now())
    updated_at = Column(TIMESTAMP(timezone=True), server_default=func.now(), onupdate=func.now())


# === PETA BULAN INDONESIA ===
BULAN_MAP = {
    "JANUARI": 1,   "FEBRUARI": 2,  "MARET": 3,
    "APRIL": 4,     "MEI": 5,       "JUNI": 6,
    "JULI": 7,      "AGUSTUS": 8,   "SEPTEMBER": 9,
    "OKTOBER": 10,  "NOVEMBER": 11, "DESEMBER": 12,
}


def ekstrak_tanggal_dari_nama_file(nama_file: str):
    """
    Ekstrak tanggal dari nama file.
    Contoh: 'LAPORAN UNSPEC HD 01 FEBRUARI 2026.xlsx' -> date(2026, 2, 1)
    """
    try:
        pola = r"HD\s+(\d{1,2})\s+(\w+)\s+(\d{4})"
        cocok = re.search(pola, nama_file, re.IGNORECASE)
        if not cocok:
            return None

        tanggal = int(cocok.group(1))
        bulan_str = cocok.group(2).upper()
        tahun = int(cocok.group(3))

        bulan = BULAN_MAP.get(bulan_str)
        if not bulan:
            print(f"  [!] Bulan '{bulan_str}' tidak dikenali")
            return None

        return date(tahun, bulan, tanggal)
    except Exception as e:
        print(f"  [!] Gagal parse tanggal: {e}")
        return None


def cari_sheet_hi(workbook):
    """Cari sheet 'data unspec HI' atau yang serupa"""
    for nama_sheet in workbook.sheetnames:
        nama_upper = nama_sheet.upper()
        if "HI" in nama_upper or "UNSPEC" in nama_upper:
            return nama_sheet

    # Fallback: sheet pertama yang bukan KURMA/Sheet1
    for nama_sheet in workbook.sheetnames:
        if nama_sheet.upper() not in ["KURMA", "SHEET1"]:
            return nama_sheet
    return None


def hitung_saldo_dari_sheet(file_path: str):
    """
    Baca sheet 'data unspec HI', hitung total saldo dan close.

    Struktur:
    - Row 5: header (No, TIME, GANDA, SEKTOR, ..., STATUS KURMA, KET)
    - Row 6+: data
    - STATUS KURMA: 'CLOSED' atau 'SISA TIKET'
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        nama_sheet = cari_sheet_hi(wb)
        if not nama_sheet:
            print(f"  [!] Sheet HI tidak ditemukan")
            wb.close()
            return None

        ws = wb[nama_sheet]

        # Cari baris header
        baris_header = None
        header_cols = {}

        for baris_idx, baris in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
            for sel in baris:
                if sel.value and str(sel.value).strip().upper() == "NO":
                    baris_header = baris_idx
                    for col_sel in baris:
                        if col_sel.value:
                            header_cols[str(col_sel.value).strip().upper()] = col_sel.column - 1
                    break
            if baris_header:
                break

        if not baris_header:
            print(f"  [!] Header row tidak ditemukan")
            wb.close()
            return None

        # Cari kolom KET (berisi CLOSED/SISA TIKET)
        # Note: STATUS KURMA = SPEC/UNSPEC, KET = CLOSED/SISA TIKET
        idx_ket = header_cols.get("KET")

        if idx_ket is None:
            # Fallback: scan data baris pertama cari kolom CLOSED/SISA TIKET
            for baris in ws.iter_rows(min_row=baris_header + 1, max_row=baris_header + 1, values_only=False):
                for sel in baris:
                    val = str(sel.value).upper() if sel.value else ""
                    if val in ["CLOSED", "SISA TIKET"]:
                        idx_ket = sel.column - 1
                        break
                break

        if idx_ket is None:
            print(f"  [!] Kolom KET (CLOSED/SISA TIKET) tidak ditemukan")
            wb.close()
            return None

        # Hitung data baris per baris
        total_saldo = 0
        total_close = 0

        for baris in ws.iter_rows(min_row=baris_header + 1, values_only=True):
            # Skip baris kosong
            if not baris[0] or str(baris[0]).strip() == "":
                continue

            # Validasi: kolom No harus angka
            try:
                int(str(baris[0]).strip())
            except ValueError:
                continue

            total_saldo += 1

            # Cek kolom KET untuk status close
            ket_val = str(baris[idx_ket]).strip().upper() if len(baris) > idx_ket and baris[idx_ket] else ""
            if ket_val == "CLOSED":
                total_close += 1

        wb.close()

        return {
            "total_saldo": total_saldo,
            "close": total_close,
            "saldo_lama": total_saldo - total_close,
        }

    except Exception as e:
        print(f"  [X] Error baca file: {e}")
        return None


def main():
    """Fungsi utama"""

    # Cek dry-run mode
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print("  IMPORT DATA HISTORIS REPORT SALDO UNSPEC")
    if dry_run:
        print("  MODE: DRY-RUN (tanpa database)")
    print("=" * 60)

    session = None

    if not dry_run:
        # Setup database
        try:
            engine = create_engine(DATABASE_URL)
            Session = sessionmaker(bind=engine)
            session = Session()
            session.execute(text("SELECT 1"))
            print("[OK] Koneksi database berhasil\n")
        except Exception as e:
            print(f"[GAGAL] Koneksi database: {e}")
            sys.exit(1)
    else:
        print("[INFO] Dry-run: skip koneksi database\n")

    # Kumpulkan file Excel
    daftar_file = []

    for nama_folder, path_folder in DATA_FOLDERS.items():
        if not path_folder.exists():
            print(f"[!] Folder '{nama_folder}' tidak ditemukan: {path_folder}")
            continue

        file_xlsx = sorted([f for f in path_folder.glob("*.xlsx") if not f.name.startswith("~$")])
        print(f"[FOLDER] {nama_folder}: {len(file_xlsx)} file")
        daftar_file.extend(file_xlsx)

    if not daftar_file:
        print("\n[GAGAL] Tidak ada file Excel!")
        if session:
            session.close()
        sys.exit(1)

    print(f"\nTotal file: {len(daftar_file)}")
    print("-" * 60)

    # Proses file
    berhasil = 0
    gagal = 0
    dilewati = 0
    hasil_semua = []

    for file_path in daftar_file:
        nama_file = file_path.name
        print(f"\n[FILE] {nama_file}")

        # Ekstrak tanggal
        tanggal = ekstrak_tanggal_dari_nama_file(nama_file)
        if not tanggal:
            print(f"  [!] Skip - tanggal tidak bisa diparse")
            dilewati += 1
            continue

        print(f"  Tanggal : {tanggal.isoformat()}")

        # Hitung saldo
        hasil = hitung_saldo_dari_sheet(str(file_path))
        if not hasil:
            gagal += 1
            continue

        print(f"  Total   : {hasil['total_saldo']}")
        print(f"  Close   : {hasil['close']}")
        print(f"  Saldo   : {hasil['saldo_lama']}")

        hasil_semua.append({
            "tanggal": tanggal,
            **hasil
        })

        if dry_run:
            berhasil += 1
            continue

        # Insert/Update ke database
        try:
            report = session.query(DailyReportDB).filter(
                DailyReportDB.date == tanggal
            ).first()

            if report:
                report.total_saldo = hasil["total_saldo"]
                report.close = hasil["close"]
                report.saldo_lama = hasil["saldo_lama"]
                report.target = TARGET_DEFAULT
                print(f"  [UPDATE] Report di-update")
            else:
                report = DailyReportDB(
                    date=tanggal,
                    total_saldo=hasil["total_saldo"],
                    close=hasil["close"],
                    saldo_lama=hasil["saldo_lama"],
                    target=TARGET_DEFAULT,
                )
                session.add(report)
                print(f"  [INSERT] Report baru")

            session.commit()
            berhasil += 1

        except Exception as e:
            session.rollback()
            print(f"  [X] Error database: {e}")
            gagal += 1

    # Ringkasan
    print("\n" + "=" * 60)
    print("  RINGKASAN IMPORT")
    print("=" * 60)
    print(f"  Berhasil  : {berhasil}")
    print(f"  Gagal     : {gagal}")
    print(f"  Dilewati  : {dilewati}")
    print(f"  Total     : {len(daftar_file)}")

    # Tampilkan hasil parsing
    if hasil_semua:
        print("\n" + "-" * 60)
        print(f"  {'Tanggal':<14} {'Total':>6} {'Close':>6} {'Saldo':>6}")
        print(f"  {'-'*14} {'-'*6} {'-'*6} {'-'*6}")
        for h in sorted(hasil_semua, key=lambda x: x["tanggal"]):
            print(f"  {h['tanggal'].isoformat():<14} {h['total_saldo']:>6} {h['close']:>6} {h['saldo_lama']:>6}")

    if session:
        session.close()

    print("\n[SELESAI]")


if __name__ == "__main__":
    main()
